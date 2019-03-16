import openpyxl
import pickle
import os
from tqdm import tqdm
from openpyxl.styles import Font, Alignment

def main():
    results_conversion = {"tile":2,"地址":3,"分类号":4,"申请号":5,"申请人":9,"专利权人":9,"发明人":10,"申请日":12,"abstract":13,"申请公布日":15,"授权公告日":15}
    
    patent_class = 'publish'
    # patent_class = 'authorization'

    excelfile='C:\\Files\\Documents\\apollo项目组\\国防科工局成果转化目录\\专利信息爬取_' + patent_class + '.xlsx'
    pklfile_2 = 'results\\' + patent_class + '\\' + patent_class + '_2.pkl'
    pklfile_filter = 'results\\' + patent_class + '\\' + patent_class + '_filter.pkl'

    with open(pklfile_2, 'rb') as f:
        results = pickle.load(f)

    # create a new excel file
    wb = openpyxl.Workbook()
    sheet = wb.get_sheet_by_name('Sheet')
    column = ('序号','发明名称','地址','分类号','申请号','专利类型','技术领域','应用领域','申请人/专利权人','发明人','法律状态','申请日','摘要','转化方式','申请/授权公布日','解密公告日','发布时间','数据来源')

    boldFont = Font(bold=True)
    centerAlignment = Alignment(horizontal="center", vertical="center")

    # 表格首行字体与对齐设置
    for columnNum in range(len(column)): # skip the first row
        sheet.cell(row=1, column=columnNum+1).value = column[columnNum]
        sheet.cell(row=1, column=columnNum+1).font = boldFont
        sheet.cell(row=1, column=columnNum+1).alignment = centerAlignment

    # 冻结行1
    sheet.freeze_panes = 'A2'

    init_row = 2
    for result in tqdm(results):
        for num in range(result['page_size']):
            page = result['patent'][num+1]
            for patent in page:
                sheet.cell(row=init_row, column=1).value = init_row-1
                sheet.row_dimensions[init_row].height = 300
                for k,v in results_conversion.items():
                    try:
                        sheet.cell(row=init_row, column=v).value = patent[k]             
                    except KeyError as e:
                        pass
                init_row += 1

    wb.save(excelfile)

if __name__ == "__main__":
    main()
